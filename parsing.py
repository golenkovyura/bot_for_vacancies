import bs4
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter
import asyncio
from aiohttp import ClientSession

HOST = 'https://www.aviasales.ru'
WIDHT_COLUMNS = 40

vacancies = []


async def get_data(session: ClientSession, url: str, name: str) -> None:
    """
    Основная функция по парсингу ифнормации о каждой ваканcсии,
    заполняет список vacancies словарями с данными.
    """
    async with session.get(url) as response:
        response_text = await response.text()
        soup = bs4.BeautifulSoup(response_text, 'lxml')

        description = soup.find('div', class_='vacancy__description') \
            .text.strip()

        responsibility = soup.find('div', class_='vacancy__responsibility') \
            .text.strip().replace('Что нужно будет делать:', '').capitalize()

        requirements = soup.find('div', class_='vacancy__requirements') \
            .text.strip().replace('Что мы ждем от тебя:', '').capitalize()

        addition = soup.find('div', class_='vacancy__addition') \
            .text.strip().replace('Что мы предлагаем:', '').capitalize()

        vacancies.append(
            {
                'Название вакансии': name,
                'URL': url,
                'Описание': description,
                'Что нужно будет делать': responsibility,
                'Ожидания от кандидата': requirements,
                'Условия работы': addition,
            }
        )


async def gather_data() -> None:
    """
    Функция создает список ивентов для ассинхронного
    парсинга вакансий.
    """
    print('Идет сбор данных...')
    async with ClientSession() as session:
        tasks = []

        response = await session.get(
            url='https://www.aviasales.ru/about/vacancies'
            )
        soup = bs4.BeautifulSoup(await response.text(), 'lxml')
        vacancies = soup.find_all(name='a', class_='vacancies_vacancy')
        for vacancie in vacancies:
            name = vacancie.find(name='p',
                                 class_='vacancies_vacancy__name').text
            url = HOST + vacancie.attrs.get('href')

            task = asyncio.create_task(get_data(session, url, name))
            tasks.append(task)
        await asyncio.gather(*tasks)


def start_parsing() -> None:
    """
    Функция, запускающая ассинхронный парсинг.
    """
    asyncio.get_event_loop().run_until_complete(gather_data())


def save_to_xlsx(user_id='') -> None:
    """
    Функция сохраняет полученные данные в excel-таблицу,
    форматируя шрифт, границы, расположение текста в ней. Принимает id
    пользователя, для которого нужно сохранить файл.
    """
    print('Выполняется сохранение данных...')
    wb = Workbook()
    ws = wb.active

    title = "Вакансии Aviasales"
    ws.title = title  # Название вкладки в excel-файле

    # Сохранение названий столбцов
    for column, key in zip(range(1, 7), vacancies[0].keys()):
        # Задаем ширину каждого столбца
        ws.column_dimensions[get_column_letter(column)].width = WIDHT_COLUMNS
        cell = ws.cell(row=1, column=column, value=key)
        # Делаем заголовки жирными
        cell.font = Font(bold=True)
        # Выравнимаем текст в ячейчках
        cell.alignment = Alignment(wrapText=True,
                                   horizontal='center',
                                   vertical='center',
                                   )
        # Задаем жирные границы ячеек для заголовка
        cell.border = Border(left=Side(style='thick'),
                             right=Side(style='thick'),
                             top=Side(style='thick'),
                             bottom=Side(style='thick'),
                             )

    # Сохранение данных в ячейки
    for row in range(len(vacancies)):
        for column, key in zip(range(1, 7), vacancies[0].keys()):
            cell = ws.cell(row=row + 2, column=column,
                           value=vacancies[row].get(key))
            # Выравнимаем текст в ячейчках
            cell.alignment = Alignment(wrapText=True,
                                       horizontal='center',
                                       vertical='center',
                                       )
            # Задаем границы ячеек
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'),
                                 )

    wb.save(f'vacancies for users/{title}_{user_id}.xlsx')
    print('Информация сохранена!')


def main(user_id=''):
    start_parsing()
    vacancies.sort(key=lambda name: name.get('Название вакансии')[0].upper())
    save_to_xlsx(user_id)


if __name__ == '__main__':
    main()
